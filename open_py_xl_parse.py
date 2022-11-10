#https://zetcode.com/python/openpyxl/

import numpy as np 
import openpyxl as pyxl

excel_path = ("../../../map_stuff/madison_cemeteries.xlsx")

# Load the worksheet
wb = pyxl.load_workbook(excel_path)

# Grab the active worksheet
sheet = wb.active

# Assign a variable from a cell
vc_cem = sheet['A2']
vc_loc = sheet['B2']
vc_lat = sheet.cell(row=2, column=3)
vc_long = sheet.cell(row=2, column=4)

print(vc_cem.value, '|', vc_loc.value)
print(vc_lat.value, '|', vc_long.value)

# Reading multiple values:
cells = sheet['A2': 'D19']
for c1, c2, c3, c4 in cells:
    print('Name:', c1.value, '| Location:', c2.value, '| Latitude:', c3.value, '| Longitude:', c4.value)

# Iterate by rows
#row = (2:19)